#!/usr/bin/env -S uv run --locked --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl==3.1.5"]
# ///
"""Read official Korail timetable files without login or reservation actions."""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass
from datetime import date as calendar_date
from datetime import time
from io import BytesIO
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from ktx_timetable import parse_timetable_rows
from openpyxl import load_workbook

BOARD_URL = (
    "https://www.korail.com/com/userBoard.do"
    "?schBcid=ticketTable&mode=list&page=1&schStr=KTX&bdCode=&Device=BH&Version=999999999"
)
FILE_BASE_URL = "https://www.korail.com/file/cubedata/COMMON/"
BOOKING_URL = "https://www.korail.com/ticket/search"
USER_AGENT = "k-skill/ktx-readonly (+https://github.com/NomaDamas/k-skill)"
KTX_TITLE = re.compile(
    r"(?:KTX|경부선|호남선|전라선|경전선|동해선|강릉선|중앙선|중부내륙선).*(?:시간표|시각표)"
)
EFFECTIVE_DATE = re.compile(r"(20\d{2})[.\s년]+(\d{1,2})[.\s월]+(\d{1,2})")


@dataclass(frozen=True)
class TimetableSource:
    title: str
    published_at: str
    download_url: str
    source_url: str = BOARD_URL


def fetch_json(url: str, timeout: float = 20.0) -> dict[str, Any]:
    request = Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urlopen(request, timeout=timeout) as response:
            return json.load(response)
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"Korail official timetable index unavailable: {exc}") from exc


def download_bytes(url: str, timeout: float = 30.0) -> bytes:
    request = Request(url, headers={"User-Agent": USER_AGENT, "Referer": BOOKING_URL})
    try:
        with urlopen(request, timeout=timeout) as response:
            return response.read()
    except (HTTPError, URLError, TimeoutError) as exc:
        raise RuntimeError(f"Korail official timetable file unavailable: {exc}") from exc


def timetable_candidates(payload: dict[str, Any]) -> list[TimetableSource]:
    candidates: list[TimetableSource] = []
    for item in payload.get("boardList", []):
        title = str(item.get("bdTitle", "")).strip()
        file_ids = item.get("fileId") or []
        if not KTX_TITLE.search(title) or not file_ids:
            continue
        file_id = str(file_ids[0]).lstrip("/")
        if not file_id.lower().endswith((".xlsx", ".xlsm")):
            continue
        candidates.append(
            TimetableSource(
                title=title,
                published_at=str(item.get("regdt", "")),
                download_url=FILE_BASE_URL + file_id,
            )
        )
    return candidates


def choose_latest_timetable(payload: dict[str, Any]) -> TimetableSource:
    candidates = timetable_candidates(payload)
    if not candidates:
        raise RuntimeError("Korail published no readable KTX timetable attachment")
    return max(candidates, key=lambda source: (source.published_at, source.title))


def effective_date(source: TimetableSource) -> str:
    match = EFFECTIVE_DATE.search(source.title)
    if match is None:
        return source.published_at.replace("-", "")
    year, month, day = match.groups()
    return f"{year}{int(month):02d}{int(day):02d}"


def choose_timetable_for_date(payload: dict[str, Any], date: str) -> TimetableSource:
    candidates = timetable_candidates(payload)
    applicable = [source for source in candidates if effective_date(source) <= date]
    if not applicable:
        raise RuntimeError(f"Korail published no KTX timetable applicable to {date}")
    return max(applicable, key=lambda source: (effective_date(source), source.published_at))


def load_workbook_bytes(content: bytes):
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Korail timetable workbook could not be parsed: {exc}") from exc


def search_public_timetable(
    *,
    dep: str,
    arr: str,
    date: str,
    earliest: str,
    latest: str,
    limit: int,
) -> dict[str, Any]:
    validate_date(date)
    start = validate_time(earliest)
    end = validate_time(latest)
    if start > end:
        raise ValueError("--time must not be later than --time-limit")
    source = choose_timetable_for_date(fetch_json(BOARD_URL), date)
    workbook = load_workbook_bytes(download_bytes(source.download_url))
    requested_date = calendar_date(int(date[:4]), int(date[4:6]), int(date[6:8]))
    trains: list[dict[str, str]] = []
    route_found = False
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheet_trains, sheet_route_found = parse_timetable_rows(
            worksheet.iter_rows(values_only=True),
            dep=dep,
            arr=arr,
            requested_date=requested_date,
            earliest=start,
            latest=end,
        )
        trains.extend(sheet_trains)
        route_found = route_found or sheet_route_found
    if not route_found:
        raise RuntimeError(f"Korail timetable station pair not found: {dep} -> {arr}")
    unique = {(train["train_no"], train["dep_time"], train["arr_time"]): train for train in trains}
    ordered = sorted(unique.values(), key=lambda train: (train["dep_time"], train["train_no"]))[:limit]
    return {
        "count": len(ordered),
        "trains": ordered,
        "date": date,
        "schedule_note": "공개 운행계획 기준이며 실시간 잔여석·운휴·지연 정보가 아닙니다.",
        "source": {"operator": "한국철도공사", **asdict(source)},
        "booking_url": BOOKING_URL,
    }


def validate_date(value: str) -> str:
    if not re.fullmatch(r"\d{8}", value):
        raise ValueError("date must use YYYYMMDD")
    try:
        calendar_date(int(value[:4]), int(value[4:6]), int(value[6:8]))
    except ValueError as exc:
        raise ValueError("date must use a valid YYYYMMDD value") from exc
    return value


def validate_time(value: str) -> str:
    if not re.fullmatch(r"\d{4}", value):
        raise ValueError("time must use HHMM")
    try:
        time.fromisoformat(f"{value[:2]}:{value[2:]}")
    except ValueError as exc:
        raise ValueError("time must use a valid HHMM value") from exc
    return f"{value[:2]}:{value[2:]}"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Korail KTX official timetable lookup (read-only)")
    commands = parser.add_subparsers(dest="command", required=True)
    search = commands.add_parser("search", help="search a published KTX operating timetable")
    search.add_argument("--dep", required=True)
    search.add_argument("--arr", required=True)
    search.add_argument("--date", required=True, help="YYYYMMDD")
    search.add_argument("--time", default="0000", help="earliest departure, HHMM")
    search.add_argument("--time-limit", default="2359", help="latest departure, HHMM")
    search.add_argument("--limit", type=int, default=10)
    commands.add_parser("source", help="show the current official timetable source")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        if args.command == "source":
            print(json.dumps(asdict(choose_latest_timetable(fetch_json(BOARD_URL))), ensure_ascii=False, indent=2))
            return 0
        if args.limit < 1 or args.limit > 50:
            raise ValueError("--limit must be between 1 and 50")
        result = search_public_timetable(
            dep=args.dep,
            arr=args.arr,
            date=args.date,
            earliest=args.time,
            latest=args.time_limit,
            limit=args.limit,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    except (RuntimeError, ValueError) as exc:
        parser.error(str(exc))
        return 2


if __name__ == "__main__":
    sys.exit(main())
